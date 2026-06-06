from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from django.db.models import Q, Sum, Count
from django.http import HttpResponse
from .models import Order, Client, Product, Equipment, OrderStatus, ProductionStage, Notification
from .forms import OrderForm, ClientForm, ProductForm, EquipmentForm, StageForm
from .utils import calculate_order_cost
import openpyxl
from openpyxl.styles import Font, Alignment
from datetime import datetime

def user_login(request):
    if request.user.is_authenticated:
        return redirect('dashboard')
    
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            login(request, user)
            return redirect('dashboard')
        else:
            messages.error(request, 'Неверное имя пользователя или пароль')
    
    return render(request, 'orders/login.html')

def user_logout(request):
    logout(request)
    return redirect('login')

@login_required
def dashboard(request):
    # Статистика для дашборда
    total_orders = Order.objects.count()
    active_orders = Order.objects.filter(status__name__in=['В работе', 'На согласовании']).count()
    completed_orders = Order.objects.filter(status__name='Выполнен').count()
    overdue_orders = Order.objects.filter(deadline_date__lt=datetime.now().date()).exclude(status__name='Выполнен').count()
    
    # Последние 10 заказов
    recent_orders = Order.objects.all().order_by('-created_at')[:10]
    
    # Активные уведомления для текущего пользователя
    notifications = Notification.objects.filter(user=request.user, is_sent=False)[:5]
    
    context = {
        'total_orders': total_orders,
        'active_orders': active_orders,
        'completed_orders': completed_orders,
        'overdue_orders': overdue_orders,
        'recent_orders': recent_orders,
        'notifications': notifications,
    }
    return render(request, 'orders/dashboard.html', context)

@login_required
def order_list(request):
    status_filter = request.GET.get('status', '')
    orders = Order.objects.all()
    
    if status_filter:
        orders = orders.filter(status__name=status_filter)
    
    statuses = OrderStatus.objects.all()
    return render(request, 'orders/order_list.html', {'orders': orders, 'statuses': statuses, 'current_status': status_filter})

@login_required
def order_create(request):
    if request.method == 'POST':
        form = OrderForm(request.POST)
        if form.is_valid():
            order = form.save(commit=False)
            order.user = request.user
            order.save()
            
            # Рассчитываем себестоимость
            order.total_cost = calculate_order_cost(order)
            order.save()
            
            messages.success(request, f'Заказ {order.order_number} успешно создан')
            return redirect('order_detail', pk=order.pk)
    else:
        form = OrderForm()
    
    return render(request, 'orders/order_form.html', {'form': form, 'title': 'Создание заказа'})

@login_required
def order_edit(request, pk):
    order = get_object_or_404(Order, pk=pk)
    
    if request.method == 'POST':
        form = OrderForm(request.POST, instance=order)
        if form.is_valid():
            order = form.save()
            order.total_cost = calculate_order_cost(order)
            order.save()
            messages.success(request, 'Заказ обновлён')
            return redirect('order_detail', pk=order.pk)
    else:
        form = OrderForm(instance=order)
    
    return render(request, 'orders/order_form.html', {'form': form, 'title': 'Редактирование заказа', 'order': order})

@login_required
def order_detail(request, pk):
    order = get_object_or_404(Order, pk=pk)
    stages = order.stages.all()
    
    # Расчёт прогресса выполнения
    total_stages = stages.count()
    completed_stages = stages.filter(stage_status='Выполнен').count()
    progress = int((completed_stages / total_stages) * 100) if total_stages > 0 else 0
    
    # Обработка добавления этапа
    if request.method == 'POST' and 'add_stage' in request.POST:
        stage_form = StageForm(request.POST)
        if stage_form.is_valid():
            stage = stage_form.save(commit=False)
            stage.order = order
            stage.save()
            messages.success(request, 'Этап добавлен')
            return redirect('order_detail', pk=order.pk)
    else:
        stage_form = StageForm()
    
    # Обработка изменения статуса этапа
    if request.method == 'POST' and 'update_stage' in request.POST:
        stage_id = request.POST.get('stage_id')
        new_status = request.POST.get('stage_status')
        stage = get_object_or_404(ProductionStage, pk=stage_id)
        stage.stage_status = new_status
        
        from datetime import date
        if new_status == 'Выполнен' and not stage.actual_end:
            stage.actual_end = date.today()
        elif new_status == 'В работе' and not stage.actual_start:
            stage.actual_start = date.today()
        
        stage.save()
        messages.success(request, f'Статус этапа "{stage.stage_name}" изменён на "{new_status}"')
        return redirect('order_detail', pk=order.pk)
    
    context = {
        'order': order,
        'stages': stages,
        'progress': progress,
        'stage_form': stage_form,
    }
    return render(request, 'orders/order_detail.html', context)

@login_required
def equipment_list(request):
    equipment = Equipment.objects.all()
    return render(request, 'orders/equipment_list.html', {'equipment': equipment})

@login_required
def equipment_create(request):
    if request.method == 'POST':
        form = EquipmentForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Оборудование добавлено')
            return redirect('equipment_list')
    else:
        form = EquipmentForm()
    
    return render(request, 'orders/equipment_form.html', {'form': form, 'title': 'Добавление оборудования'})

@login_required
def equipment_edit(request, pk):
    equipment = get_object_or_404(Equipment, pk=pk)
    
    if request.method == 'POST':
        form = EquipmentForm(request.POST, instance=equipment)
        if form.is_valid():
            form.save()
            messages.success(request, 'Оборудование обновлено')
            return redirect('equipment_list')
    else:
        form = EquipmentForm(instance=equipment)
    
    return render(request, 'orders/equipment_form.html', {'form': form, 'title': 'Редактирование оборудования'})

@login_required
def client_list(request):
    clients = Client.objects.all()
    return render(request, 'orders/client_list.html', {'clients': clients})

@login_required
def client_create(request):
    if request.method == 'POST':
        form = ClientForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Клиент добавлен')
            return redirect('client_list')
    else:
        form = ClientForm()
    
    return render(request, 'orders/client_form.html', {'form': form, 'title': 'Добавление клиента'})

@login_required
def client_edit(request, pk):
    client = get_object_or_404(Client, pk=pk)
    
    if request.method == 'POST':
        form = ClientForm(request.POST, instance=client)
        if form.is_valid():
            form.save()
            messages.success(request, 'Клиент обновлён')
            return redirect('client_list')
    else:
        form = ClientForm(instance=client)
    
    return render(request, 'orders/client_form.html', {'form': form, 'title': 'Редактирование клиента'})

@login_required
def product_list(request):
    products = Product.objects.all()
    return render(request, 'orders/product_list.html', {'products': products})

@login_required
def product_create(request):
    if request.method == 'POST':
        form = ProductForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Продукция добавлена')
            return redirect('product_list')
    else:
        form = ProductForm()
    
    return render(request, 'orders/product_form.html', {'form': form, 'title': 'Добавление продукции'})

@login_required
def product_edit(request, pk):
    product = get_object_or_404(Product, pk=pk)
    
    if request.method == 'POST':
        form = ProductForm(request.POST, instance=product)
        if form.is_valid():
            form.save()
            messages.success(request, 'Продукция обновлена')
            return redirect('product_list')
    else:
        form = ProductForm(instance=product)
    
    return render(request, 'orders/product_form.html', {'form': form, 'title': 'Редактирование продукции'})

@login_required
def report(request):
    # Фильтры
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    report_type = request.GET.get('report_type', 'orders')
    
    orders = Order.objects.all()
    
    if date_from:
        orders = orders.filter(created_at__date__gte=date_from)
    if date_to:
        orders = orders.filter(created_at__date__lte=date_to)
    
    # Отчёт по загрузке оборудования
    equipment_load = []
    for eq in Equipment.objects.all():
        stages = ProductionStage.objects.filter(equipment=eq)
        load_count = stages.count()
        equipment_load.append({
            'name': eq.name,
            'load_count': load_count,
            'is_available': eq.is_available,
        })
    
    context = {
        'orders': orders,
        'equipment_load': equipment_load,
        'date_from': date_from,
        'date_to': date_to,
        'report_type': report_type,
    }
    return render(request, 'orders/report.html', context)

@login_required
def export_excel(request):
    """Экспорт заказов в Excel"""
    orders = Order.objects.all()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Заказы"
    
    # Заголовки
    headers = ['№', 'Номер заказа', 'Клиент', 'Продукция', 'Тираж', 'Статус', 'Срок сдачи', 'Себестоимость', 'Дата создания']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # Данные
    for row, order in enumerate(orders, 2):
        ws.cell(row=row, column=1, value=row-1)
        ws.cell(row=row, column=2, value=order.order_number)
        ws.cell(row=row, column=3, value=order.client.name)
        ws.cell(row=row, column=4, value=order.product.name)
        ws.cell(row=row, column=5, value=order.quantity)
        ws.cell(row=row, column=6, value=order.status.name)
        ws.cell(row=row, column=7, value=str(order.deadline_date))
        ws.cell(row=row, column=8, value=float(order.total_cost))
        ws.cell(row=row, column=9, value=order.created_at.strftime('%d.%m.%Y'))
    
    # Автоширина колонок
    for col in range(1, 10):
        ws.column_dimensions[chr(64 + col)].width = 15
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=orders_{datetime.now().strftime("%Y%m%d")}.xlsx'
    wb.save(response)
    
    return response